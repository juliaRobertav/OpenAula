from openpyxl import load_workbook
wbt = load_workbook('C:/Users/sn1021328/Desktop/teste.xlsx')
wbp = load_workbook('C:/Users/sn1021328/Desktop/professores.xlsx')
p = wbp['Sheet1']
t = wbt['P3']
doutores = []

for i in range(30):
    if p.cell(row=i+1, column=3).value == 'DOUTORADO':
        doutores.append(p.cell(row=i+1, column=2).value)
for i in range(len(doutores)):
    print(doutores[i])

for i in range(len(doutores)):
    t[f'A{i+1}'] = i
for i in range(len(doutores)):
    t[f'B{i+2}'] = doutores[i]

t['A1'] = 'Nº'
t['B1'] = 'Nome'

wbt.save('C:/Users/sn1021328/Desktop/teste.xlsx')
